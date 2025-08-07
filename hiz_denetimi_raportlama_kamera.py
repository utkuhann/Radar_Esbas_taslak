import cv2
import numpy as np
import torch
import os
from datetime import datetime
from collections import defaultdict
import time
import threading
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from google.cloud import vision

# --- WEB SUNUCUSU İÇİN GEREKLİ KÜTÜPHANELER ---
from flask import Flask, Response, render_template_string

# --- PROJE KÜTÜPHANELERİ ---
from basicsr.archs.rrdbnet_arch import RRDBNet
from realesrgan import RealESRGANer
from ultralytics import YOLO
import supervision as sv

# ==============================================================================
# --- 1. MERKEZİ YAPILANDIRMA (BURAYI DÜZENLEYİN) ---
# ==============================================================================
# İşlemek istediğiniz her kamera için bir sözlük (dictionary) ekleyin.
# url: Kamera IP akışının tam adresi (RTSP, HTTP, vb.).
# line_ratio: Tespit çizgisinin ekrandaki dikey konumu (0.0 ile 1.0 arası).
# ppm: O kameradaki hız hesabı için kullanılacak Piksel Per Metre değeri.

CAMERA_CONFIGS = [
    {
        "url": "http://15.25.110.132/stream1", # Örnek URL 1 - Kendi URL'nizle değiştirin
        "line_ratio": 0.6,
        "ppm": 20
    },
    {
        "url": "http://15.25.110.132/stream2", # Örnek URL 2 - Kendi URL'nizle değiştirin
        "line_ratio": 0.55,
        "ppm": 30
    },
    {
        "url": "http://15.25.110.132/stream3", # Örnek URL 3 - Kendi URL'nizle değiştirin
        "line_ratio": 0.7,
        "ppm": 30
    },
    {
        "url": "http://15.25.110.132/stream4", # Örnek URL 4 - Kendi URL'nizle değiştirin
        "line_ratio": 0.7,
        "ppm": 30
    }
]

# --- Global Ayarlar ---
ARAC_MODEL_YOLU = "yolov8n.pt"
PLAKA_MODEL_YOLU = "lapi.pt"
REAL_ESRGAN_MODEL_YOLU = 'Real-ESRGAN/RealESRGAN_x4plus.pth'
TARGET_HEIGHT = 480
VIDEO_FPS = 30  # Kamera FPS değerine yakın bir değer girin
HIZ_LIMITI_KMH = 60
CIKTI_KLASORU = "hiz_ihlalleri_yuksek_kalite"
EXCEL_RAPOR_ADI = "birlesik_ihlal_raporu.xlsx"
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"

# ==============================================================================
# --- 2. FLASK UYGULAMASI VE GLOBAL DEĞİŞKENLER ---
# ==============================================================================
app = Flask(__name__)

# Her video akışının son karesini (frame) saklamak için bir sözlük
output_frames = {i: None for i in range(len(CAMERA_CONFIGS))}
# Bu sözlüğe erişirken oluşabilecek çakışmaları önlemek için kilit mekanizması
frame_locks = {i: threading.Lock() for i in range(len(CAMERA_CONFIGS))}
# Excel dosyasına yazma işlemini senkronize etmek için kilit
excel_lock = threading.Lock()


# Dinamik HTML Şablonu (Değişiklik yok)
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Çoklu Kamera Hız Denetimi Canlı Yayını</title>
    <style>
        body { font-family: sans-serif; text-align: center; background-color: #2c2c2c; color: #f0f0f0; margin: 0; padding: 20px; }
        h1 { color: #fff; }
        #video-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin: 20px auto; max-width: 1800px; }
        .video-container { border: 2px solid #555; background-color: #000; position: relative; }
        .video-container h3 { position: absolute; top: 5px; left: 10px; margin: 0; padding: 5px; background-color: rgba(0,0,0,0.6); border-radius: 5px; font-size: 16px; }
        .video-container img { width: 100%; height: auto; display: block; }
    </style>
</head>
<body>
    <h1>Çoklu Kamera Hız Denetimi</h1>
    <div id="video-grid">
        {% for i in range(video_count) %}
        <div class="video-container">
            <img src="{{ url_for('video_feed', video_index=i) }}" alt="Kamera Yayını {{ i+1 }}">
            <h3>Kamera {{ i+1 }}</h3>
        </div>
        {% endfor %}
    </div>
</body>
</html>
"""

# ==============================================================================
# --- 3. MODELLERİN YÜKLENMESİ VE YARDIMCI FONKSİYONLAR ---
# ==============================================================================
print(f"Kullanılan Cihaz: {DEVICE}")
print("Modeller yükleniyor...")
arac_model = YOLO(ARAC_MODEL_YOLU).to(DEVICE)
plaka_model = YOLO(PLAKA_MODEL_YOLU).to(DEVICE)
upsampler = None
try:
    esrgan_model = RRDBNet(num_in_ch=3, num_out_ch=3, num_feat=64, num_block=23, num_grow_ch=32, scale=4)
    upsampler = RealESRGANer(scale=4, model_path=REAL_ESRGAN_MODEL_YOLU, model=esrgan_model, tile=0, half=DEVICE=='cuda', gpu_id=0 if DEVICE=='cuda' else None)
    print("Real-ESRGAN modeli başarıyla yüklendi.")
except Exception as e:
    print(f"UYARI: Real-ESRGAN modeli yüklenemedi: {e}")

def metni_oku(image_np: np.ndarray) -> str | None:
    try:
        if image_np is None: return None
        client = vision.ImageAnnotatorClient()
        success, encoded_image = cv2.imencode('.png', image_np)
        if not success: return None
        content = encoded_image.tobytes()
        image = vision.Image(content=content)
        response = client.text_detection(image=image)
        if response.error.message: return None
        if response.text_annotations:
            return "".join(response.text_annotations[0].description.split())
    except Exception as e:
        print(f"OCR Hatası: {e}")
    return None

def iyilestir_ve_plaka_oku(arac_resmi: np.ndarray, plaka_model: YOLO, upsampler: RealESRGANer) -> tuple[str | None, np.ndarray | None]:
    plaka_tespitleri = plaka_model(arac_resmi, verbose=False)[0]
    if len(plaka_tespitleri.boxes) == 0: return None, None
    plaka_xyxy = plaka_tespitleri.boxes[0].xyxy[0].cpu().numpy()
    plaka_kirpma = sv.crop_image(image=arac_resmi, xyxy=plaka_xyxy)
    if plaka_kirpma.size == 0: return None, None
    h, w, _ = plaka_kirpma.shape
    buyutulmus_plaka = cv2.resize(plaka_kirpma, (w*3, h*3), interpolation=cv2.INTER_CUBIC)
    gray_plate = cv2.cvtColor(buyutulmus_plaka, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    contrast_plate = clahe.apply(gray_plate)
    plaka_metni = metni_oku(contrast_plate)
    final_enhanced_plate = buyutulmus_plaka
    if not plaka_metni or len(plaka_metni) < 5:
        if upsampler:
            try:
                super_res_plaka, _ = upsampler.enhance(plaka_kirpma, outscale=4)
                gray_sr = cv2.cvtColor(super_res_plaka, cv2.COLOR_BGR2GRAY)
                clahe_sr = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                contrast_sr = clahe_sr.apply(gray_sr)
                plaka_metni_sr = metni_oku(contrast_sr)
                if plaka_metni_sr and len(plaka_metni_sr) > len(plaka_metni or ""):
                    plaka_metni = plaka_metni_sr
                final_enhanced_plate = super_res_plaka
            except Exception as e:
                print(f"Real-ESRGAN hatası: {e}")
    return plaka_metni, final_enhanced_plate

def append_to_excel_report(veri_kaydi: dict, dosya_yolu: str):
    """
    Tek bir ihlal kaydını Excel dosyasına ekler.
    Dosya yoksa oluşturur, varsa yeni satır ekler. Thread-safe'dir.
    """
    with excel_lock:
        try:
            df = pd.DataFrame([veri_kaydi])
            file_exists = os.path.isfile(dosya_yolu)

            if not file_exists:
                # Sütun adlarını daha okunabilir hale getir ve sırala
                new_cols = {'kamera_id': 'Kamera ID', 'timestamp': 'Tarih/Saat', 'tracker_id': 'Takip ID', 'hiz_kmh': 'Hız (km/h)', 'plaka_metni': 'Plaka Metni', 'arac_resim_yolu': 'Araç Resim Yolu', 'plaka_resim_yolu': 'Plaka Resim Yolu'}
                df.rename(columns=new_cols, inplace=True)
                df.to_excel(dosya_yolu, index=False, engine='openpyxl')
                
                # Yeni dosyanın sütun genişliklerini ayarla
                wb = load_workbook(dosya_yolu)
                ws = wb.active
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 50
                ws.column_dimensions['G'].width = 50
                ws.row_dimensions[1].height = 20
                wb.save(dosya_yolu)

            else:
                # Mevcut dosyaya başlık olmadan ekle
                with pd.ExcelWriter(dosya_yolu, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

            print(f"Kamera {veri_kaydi['kamera_id']} ihlali başarıyla rapora eklendi.")

        except Exception as e:
            print(f"Excel raporuna ekleme yapılırken hata oluştu: {e}")


# ==============================================================================
# --- 4. ANA İŞLEME FONKSİYONU (HER KAMERA THREAD'İ İÇİN) ---
# ==============================================================================
def process_video_stream(video_index, config):
    """
    Tek bir IP kamera akışını işler. Bu fonksiyon her kamera için ayrı bir thread'de çalışır.
    """
    global output_frames

    # Her thread kendi tracker ve diğer durum değişkenlerine sahip olur.
    tracker = sv.ByteTrack(frame_rate=VIDEO_FPS, track_activation_threshold=0.25)
    box_annotator = sv.BoxAnnotator(thickness=2)
    track_gecmisi = defaultdict(list)
    best_shot_info = {}
    processed_violators = set()
    
    camera_url = config['url']
    tespit_cizgisi_orani = config['line_ratio']
    piksel_per_metre = config['ppm']

    cap = cv2.VideoCapture(camera_url)
    if not cap.isOpened():
        print(f"HATA: Kamera {video_index+1} için IP akışı açılamadı: {camera_url}")
        # Hata durumunda web arayüzünde bir hata mesajı göster
        error_frame = np.zeros((TARGET_HEIGHT, int(TARGET_HEIGHT * 16/9), 3), dtype=np.uint8)
        cv2.putText(error_frame, f"Kamera {video_index+1}: Baglanti Hatasi", (50, int(TARGET_HEIGHT/2)), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
        with frame_locks[video_index]:
            ret, buffer = cv2.imencode('.jpg', error_frame)
            output_frames[video_index] = buffer.tobytes()
        return

    print(f"Kamera {video_index+1} ({camera_url}) işlemi başlıyor...")

    while True:
        ret, frame = cap.read()
        if not ret:
            print(f"UYARI: Kamera {video_index+1} akışından kare okunamadı. Tekrar deneniyor...")
            time.sleep(2) # Tekrar denemeden önce kısa bir bekleme
            cap.release()
            cap = cv2.VideoCapture(camera_url) # Bağlantıyı yeniden kurmayı dene
            continue

        original_height, original_width, _ = frame.shape
        low_res_frame = cv2.resize(frame, (int(original_width * TARGET_HEIGHT / original_height), TARGET_HEIGHT))
        H, W, _ = low_res_frame.shape
        
        results = arac_model(low_res_frame, verbose=False, conf=0.4)[0]
        detections = sv.Detections.from_ultralytics(results)
        detections = tracker.update_with_detections(detections)
        VEHICLE_CLASSES = [2, 3, 5, 7] # araba, motosiklet, otobüs, kamyon
        detections = detections[np.isin(detections.class_id, VEHICLE_CLASSES)]
        
        for i in range(len(detections)):
            xyxy = detections.xyxy[i]
            tracker_id = detections.tracker_id[i]
            # En net fotoğrafı (en büyük alanı) sakla
            xyxy_orig = xyxy * np.array([original_width / W, original_height / H, original_width / W, original_height / H])
            arac_alani = (xyxy_orig[2] - xyxy_orig[0]) * (xyxy_orig[3] - xyxy_orig[1])
            if tracker_id not in best_shot_info or arac_alani > best_shot_info[tracker_id]['alan']:
                best_shot_info[tracker_id] = {'frame': frame.copy(), 'xyxy': xyxy_orig, 'alan': arac_alani}

            # Hız hesaplama
            center_y = int(xyxy_orig[3]) # Aracın alt orta noktasını kullan
            track_gecmisi[tracker_id].append(center_y)

            if len(track_gecmisi[tracker_id]) > 1 and tracker_id not in processed_violators:
                onceki_y = track_gecmisi[tracker_id][-2]
                tespit_cizgisi_y_orig = int(original_height * tespit_cizgisi_orani)
                
                # Çizgiyi geçiş anını tespit et
                if (onceki_y < tespit_cizgisi_y_orig <= center_y) or (onceki_y > tespit_cizgisi_y_orig >= center_y):
                    hiz_kmh = (abs(center_y - onceki_y) / piksel_per_metre / (1 / VIDEO_FPS)) * 3.6
                    if hiz_kmh > HIZ_LIMITI_KMH:
                        print(f"--- KAMERA {video_index+1}: ID {tracker_id} İHLAL TESPİT EDİLDİ ({int(hiz_kmh)} km/h) ---")
                        processed_violators.add(tracker_id)
                        
                        # --- İHLALİ ANINDA İŞLE VE KAYDET ---
                        best_info = best_shot_info.get(tracker_id)
                        if best_info:
                            arac_resmi = sv.crop_image(image=best_info['frame'], xyxy=best_info['xyxy'])
                            plaka_metni, iyilestirilmis_plaka_resmi = iyilestir_ve_plaka_oku(arac_resmi, plaka_model, upsampler)
                            
                            timestamp = datetime.now()
                            timestamp_str = timestamp.strftime("%Y%m%d_%H%M%S")
                            arac_resim_yolu = os.path.join(CIKTI_KLASORU, f"cam{video_index+1}_arac_{tracker_id}_{timestamp_str}.jpg")
                            plaka_resim_yolu = os.path.join(CIKTI_KLASORU, f"cam{video_index+1}_plaka_{tracker_id}_{timestamp_str}.jpg")
                            
                            cv2.imwrite(arac_resim_yolu, arac_resmi)
                            if iyilestirilmis_plaka_resmi is not None:
                                cv2.imwrite(plaka_resim_yolu, iyilestirilmis_plaka_resmi)
                            else:
                                plaka_resim_yolu = None # Yolu None yap
                            
                            # Excel'e eklenecek veri kaydını oluştur
                            ihlal_kaydi = {
                                "kamera_id": video_index + 1,
                                "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                                "tracker_id": tracker_id, "hiz_kmh": int(hiz_kmh),
                                "plaka_metni": plaka_metni or "OKUNAMADI",
                                "arac_resim_yolu": arac_resim_yolu, 
                                "plaka_resim_yolu": plaka_resim_yolu or "YOK",
                            }
                            # Excel'e ekleme fonksiyonunu çağır
                            append_to_excel_report(ihlal_kaydi, EXCEL_RAPOR_ADI)

        # Web arayüzü için kareyi hazırla
        annotated_frame = box_annotator.annotate(scene=low_res_frame.copy(), detections=detections)
        tespit_cizgisi_y_low_res = int(H * tespit_cizgisi_orani)
        cv2.line(annotated_frame, (0, tespit_cizgisi_y_low_res), (W, tespit_cizgisi_y_low_res), (0, 255, 0), 2)
        
        with frame_locks[video_index]:
            ret, buffer = cv2.imencode('.jpg', annotated_frame)
            output_frames[video_index] = buffer.tobytes()

    cap.release()

def generate_frame_for_web(video_index):
    """
    Belirli bir video indeksi için web akışını oluşturur.
    Sadece global output_frames değişkeninden kareleri okur ve yayınlar.
    """
    while True:
        with frame_locks[video_index]:
            frame = output_frames[video_index]
        
        if frame is None:
            time.sleep(0.05)
            continue
        
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
        
        time.sleep(1 / VIDEO_FPS)

# ==============================================================================
# --- 5. FLASK ROUTE'LARI VE ANA ÇALIŞTIRMA BLOĞU ---
# ==============================================================================
@app.route('/')
def index():
    """Ana sayfayı, kamera sayısı kadar gösterim alanı ile oluşturur."""
    return render_template_string(HTML_TEMPLATE, video_count=len(CAMERA_CONFIGS))

@app.route('/video_feed/<int:video_index>')
def video_feed(video_index):
    """Belirli bir kamera akışını yayınlar."""
    if video_index >= len(CAMERA_CONFIGS):
        return "Geçersiz kamera indeksi", 404
    return Response(generate_frame_for_web(video_index),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

if __name__ == "__main__":
    os.makedirs(CIKTI_KLASORU, exist_ok=True)
    
    # Her kamera için bir işleme thread'i oluştur ve başlat
    for i, config in enumerate(CAMERA_CONFIGS):
        thread = threading.Thread(target=process_video_stream, args=(i, config))
        thread.daemon = True # Ana program kapandığında thread'lerin de kapanmasını sağlar
        thread.start()

    # Flask web sunucusunu başlat
    print("\nWeb sunucusu başlatılıyor...")
    print(f"Arayüze erişmek için tarayıcınızda http://127.0.0.1:5001 adresini açın.")
    # Not: production ortamında debug=False kullanılmalıdır.
    app.run(host='0.0.0.0', port=5001, debug=False, threaded=True)