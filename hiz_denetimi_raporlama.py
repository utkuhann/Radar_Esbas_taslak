import cv2
import numpy as np
import torch
import os
from datetime import datetime
from collections import defaultdict
import time
import threading

# --- WEB SUNUCUSU İÇİN GEREKLİ KÜTÜPHANELER ---
from flask import Flask, Response, render_template_string

# --- PROJE KÜTÜPHANELERİ ---
from basicsr.archs.rrdbnet_arch import RRDBNet
from realesrgan import RealESRGANer
from ultralytics import YOLO
import supervision as sv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from google.cloud import vision

# ==============================================================================
# --- 1. MERKEZİ YAPILANDIRMA (BURAYI DÜZENLEYİN) ---
# ==============================================================================
# Her video için üst ve alt çizgiyi ve aralarındaki mesafeyi ayarlayın.
# path: Video dosyasının tam yolu.
# line_top_ratio: Üst tespit çizgisinin konumu (0.0 - 1.0 arası).
# line_bottom_ratio: Alt tespit çizgisinin konumu (0.0 - 1.0 arası).
# distance_m: İki çizgi arasındaki GERÇEK MESAFE (metre cinsinden). DOĞRU HESAPLAMA İÇİN KRİTİKTİR.

VIDEO_CONFIGS = [
    {
        "path": "/Users/utkuhanergene/Documents/SoftwareProjects/MyVSCode/supervision/examples/speed_estimation/plaka_tanima_projesi/videos/traffic1.mp4",
        "line_top_ratio": 0.45,
        "line_bottom_ratio": 0.75,
        "distance_m": 15  # Bu iki çizgi arası gerçekte 15 metre varsayıldı
    },
    {
        "path": "/Users/utkuhanergene/Documents/SoftwareProjects/MyVSCode/supervision/examples/speed_estimation/plaka_tanima_projesi/videos/traffic2.mp4",
        "line_top_ratio": 0.40,
        "line_bottom_ratio": 0.80,
        "distance_m": 20  # Bu iki çizgi arası gerçekte 20 metre varsayıldı
    },
    {
        "path": "/Users/utkuhanergene/Documents/SoftwareProjects/MyVSCode/supervision/examples/speed_estimation/plaka_tanima_projesi/videos/traffic3.mp4",
        "line_top_ratio": 0.50,
        "line_bottom_ratio": 0.70,
        "distance_m": 10
    },
    {
        "path": "/Users/utkuhanergene/Documents/SoftwareProjects/MyVSCode/supervision/examples/speed_estimation/plaka_tanima_projesi/videos/traffic4.mp4",
        "line_top_ratio": 0.35,
        "line_bottom_ratio": 0.65,
        "distance_m": 12
    }
]

# --- Global Ayarlar ---
ARAC_MODEL_YOLU = "yolov8n.pt"
PLAKA_MODEL_YOLU = "lapi.pt"
REAL_ESRGAN_MODEL_YOLU = 'Real-ESRGAN/RealESRGAN_x4plus.pth'
TARGET_HEIGHT = 480
VIDEO_FPS = 30 # Not: Bu değer artık anlık hız hesabında kullanılmıyor ama tracker için hala önemli.
HIZ_LIMITI_KMH = 60
CIKTI_KLASORU = "hiz_ihlalleri_yuksek_kalite"
EXCEL_RAPOR_ADI = "birlesik_ihlal_raporu_dogru_hesap.xlsx"
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"

# Diğer global değişkenler ve Flask uygulaması önceki kod ile aynı
# ... (HTML_TEMPLATE, app, output_frames, locks vb. burada) ...
app = Flask(__name__)
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Çoklu Hız Denetimi Yayını (Doğru Hesaplama)</title>
    <style>
        body { font-family: sans-serif; text-align: center; background-color: #2c2c2c; color: #f0f0f0; margin: 0; padding: 20px; }
        h1 { color: #fff; }
        #video-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin: 20px auto; max-width: 1800px; }
        .video-container { border: 2px solid #555; background-color: #000; position: relative; }
        .video-container h3 { position: absolute; top: 5px; left: 10px; margin: 0; padding: 5px; background-color: rgba(0,0,0,0.6); border-radius: 5px; font-size: 16px; }
        .video-container img { width: 100%; height: auto; display: block; }
    </style>
</head>
<body>
    <h1>Çoklu Kamera Hız Denetimi - Ortalama Hız Hesaplama</h1>
    <div id="video-grid">
        {% for i in range(video_count) %}
        <div class="video-container">
            <img src="{{ url_for('video_feed', video_index=i) }}" alt="Video Stream {{ i+1 }}">
            <h3>Kamera {{ i+1 }}</h3>
        </div>
        {% endfor %}
    </div>
</body>
</html>
"""
output_frames = {i: None for i in range(len(VIDEO_CONFIGS))}
frame_locks = {i: threading.Lock() for i in range(len(VIDEO_CONFIGS))}
all_ihlal_kayitlari = []
all_kayitlari_lock = threading.Lock()
processing_threads = []

# ==============================================================================
# --- 3. MODELLERİN YÜKLENMESİ VE YARDIMCI FONKSİYONLAR ---
# ==============================================================================
print(f"Kullanılan Cihaz: {DEVICE}")
print("Modeller yükleniyor...")
arac_model = YOLO(ARAC_MODEL_YOLU).to(DEVICE)
plaka_model = YOLO(PLAKA_MODEL_YOLU).to(DEVICE)
upsampler = None
try:
    esrgan_model = RRDBNet(num_in_ch=3, num_out_ch=3, num_feat=64, num_block=23, num_grow_ch=32, scale=4)
    upsampler = RealESRGANer(scale=4, model_path=REAL_ESRGAN_MODEL_YOLU, model=esrgan_model, tile=0, half=DEVICE=='cuda', gpu_id=0 if DEVICE=='cuda' else None)
    print("Real-ESRGAN modeli başarıyla yüklendi.")
except Exception as e:
    print(f"UYARI: Real-ESRGAN modeli yüklenemedi: {e}")

# Yardımcı fonksiyonlar (metni_oku, iyilestir_ve_plaka_oku, excel_raporu_olustur)
# önceki kod ile tamamen aynıdır, değişiklik yoktur.
# ... (Bu fonksiyonlar buraya eklenecek) ...
def metni_oku(image_np: np.ndarray) -> str | None:
    try:
        if image_np is None: return None
        client = vision.ImageAnnotatorClient()
        success, encoded_image = cv2.imencode('.png', image_np)
        if not success: return None
        content = encoded_image.tobytes()
        image = vision.Image(content=content)
        response = client.text_detection(image=image)
        if response.error.message: return None
        if response.text_annotations:
            return "".join(response.text_annotations[0].description.split())
    except Exception as e:
        print(f"OCR Hatası: {e}")
    return None

def iyilestir_ve_plaka_oku(arac_resmi: np.ndarray, plaka_model: YOLO, upsampler: RealESRGANer) -> tuple[str | None, np.ndarray | None]:
    plaka_tespitleri = plaka_model(arac_resmi, verbose=False)[0]
    if len(plaka_tespitleri.boxes) == 0: return None, None
    plaka_xyxy = plaka_tespitleri.boxes[0].xyxy[0].cpu().numpy()
    plaka_kirpma = sv.crop_image(image=arac_resmi, xyxy=plaka_xyxy)
    if plaka_kirpma.size == 0: return None, None
    h, w, _ = plaka_kirpma.shape
    buyutulmus_plaka = cv2.resize(plaka_kirpma, (w*3, h*3), interpolation=cv2.INTER_CUBIC)
    gray_plate = cv2.cvtColor(buyutulmus_plaka, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    contrast_plate = clahe.apply(gray_plate)
    plaka_metni = metni_oku(contrast_plate)
    final_enhanced_plate = buyutulmus_plaka
    if not plaka_metni or len(plaka_metni) < 5:
        if upsampler:
            try:
                super_res_plaka, _ = upsampler.enhance(plaka_kirpma, outscale=4)
                gray_sr = cv2.cvtColor(super_res_plaka, cv2.COLOR_BGR2GRAY)
                clahe_sr = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                contrast_sr = clahe_sr.apply(gray_sr)
                plaka_metni_sr = metni_oku(contrast_sr)
                if plaka_metni_sr and len(plaka_metni_sr) > len(plaka_metni or ""):
                    plaka_metni = plaka_metni_sr
                final_enhanced_plate = super_res_plaka
            except Exception as e:
                print(f"Real-ESRGAN hatası: {e}")
    return plaka_metni, final_enhanced_plate

def excel_raporu_olustur(veri_listesi: list, dosya_yolu: str):
    if not veri_listesi:
        print("Raporlanacak toplam ihlal bulunamadı.")
        return
    df = pd.DataFrame(veri_listesi)
    df = df.sort_values(by=['kamera_id', 'timestamp']).reset_index(drop=True)
    ws_col_widths = {'A': 20, 'B': 12, 'C': 12, 'D': 15, 'E': 20, 'F': 50, 'G': 50, 'H': 30, 'I': 25}
    new_cols = {'kamera_id': 'Kamera ID', 'timestamp': 'Tarih/Saat', 'tracker_id': 'Takip ID', 'hiz_kmh': 'Hız (km/h)', 'plaka_metni': 'Plaka Metni', 'arac_resim_yolu': 'Araç Resim Yolu', 'plaka_resim_yolu': 'Plaka Resim Yolu', 'H': "Arac Goruntusu", 'I': "Plaka Goruntusu"}
    df.rename(columns=new_cols, inplace=True)
    df.to_excel(dosya_yolu, index=False, engine='openpyxl')
    print(f"Tüm veriler '{dosya_yolu}' dosyasına kaydedildi. Resimler ekleniyor...")
    try:
        wb = load_workbook(dosya_yolu)
        ws = wb.active
        for col, width in ws_col_widths.items(): ws.column_dimensions[col].width = width
        ws['H1'] = new_cols['H']
        ws['I1'] = new_cols['I']
        for i, kayit in enumerate(df.to_dict('records'), start=2):
            ws.row_dimensions[i].height = 120
            arac_resim_yolu = kayit.get(new_cols['arac_resim_yolu'])
            plaka_resim_yolu = kayit.get(new_cols['plaka_resim_yolu'])
            if arac_resim_yolu and os.path.exists(arac_resim_yolu):
                img = OpenpyxlImage(arac_resim_yolu); img.height, img.width = 150, 200; ws.add_image(img, f'H{i}')
            if plaka_resim_yolu and os.path.exists(plaka_resim_yolu):
                img = OpenpyxlImage(plaka_resim_yolu); img.height, img.width = 100, 150; ws.add_image(img, f'I{i}')
        wb.save(dosya_yolu)
        print(f"Resimler başarıyla birleşik Excel raporuna eklendi.")
    except Exception as e:
        print(f"Excel'e resim eklenirken bir hata oluştu: {e}")

# ==============================================================================
# --- 4. ANA İŞLEME FONKSİYONU (İKİ ÇİZGİLİ SİSTEME GÜNCELLENDİ) ---
# ==============================================================================
def process_video_stream(video_index, config):
    global output_frames, all_ihlal_kayitlari

    # Her thread kendi durum değişkenlerine sahip.
    tracker = sv.ByteTrack(frame_rate=VIDEO_FPS, track_activation_threshold=0.25)
    box_annotator = sv.BoxAnnotator(thickness=2)
    track_gecmisi = defaultdict(list)
    best_shot_info = {}
    final_violators = []

    # --- YENİ: İKİ ÇİZGİLİ SİSTEM İÇİN DURUM DEĞİŞKENLERİ ---
    # Üst çizgiyi geçen araçların giriş zamanlarını saklamak için sözlük
    entry_times = {} # {tracker_id: entry_time}

    # Yapılandırmadan ilgili ayarları al
    video_path = config['path']
    line_top_ratio = config['line_top_ratio']
    line_bottom_ratio = config['line_bottom_ratio']
    distance_m = config['distance_m']

    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print(f"HATA: Kamera {video_index+1} için video dosyası açılamadı: {video_path}")
        return

    print(f"Kamera {video_index+1} ({os.path.basename(video_path)}) işlemi başlıyor...")

    while True:
        ret, frame = cap.read()
        if not ret: break

        original_height, original_width, _ = frame.shape
        low_res_frame = cv2.resize(frame, (int(original_width * TARGET_HEIGHT / original_height), TARGET_HEIGHT))
        H, W, _ = low_res_frame.shape
        
        results = arac_model(low_res_frame, verbose=False, conf=0.4)[0]
        detections = sv.Detections.from_ultralytics(results)
        detections = tracker.update_with_detections(detections)
        VEHICLE_CLASSES = [2, 3, 5, 7]
        detections = detections[np.isin(detections.class_id, VEHICLE_CLASSES)]
        
        # --- DEĞİŞTİ: HESAPLAMA MANTIĞI GÜNCELLENDİ ---
        # Çizgilerin orijinal çözünürlükteki Y koordinatları
        line_top_y = int(original_height * line_top_ratio)
        line_bottom_y = int(original_height * line_bottom_ratio)

        for i in range(len(detections)):
            xyxy = detections.xyxy[i]
            tracker_id = detections.tracker_id[i]
            
            # Aracın alt-orta noktasını referans alalım
            center_x = int((xyxy[0] + xyxy[2]) / 2)
            center_y = int(xyxy[3]) # Bounding box'ın alt çizgisi
            
            # Orijinal çözünürlükteki koordinat
            center_y_orig = int(center_y * original_height / H)

            # Geçmiş takibini güncelle
            track_gecmisi[tracker_id].append(center_y_orig)
            if len(track_gecmisi[tracker_id]) < 2:
                continue
            
            previous_y = track_gecmisi[tracker_id][-2]

            # 1. Adım: Araç üst çizgiyi geçti mi?
            if previous_y < line_top_y <= center_y_orig and tracker_id not in entry_times:
                entry_times[tracker_id] = time.time()
                print(f"Kamera {video_index+1}: ID {tracker_id} GİRİŞ YAPTI.")

            # 2. Adım: Araç alt çizgiyi geçti mi? (VE DAHA ÖNCE GİRİŞ YAPMIŞ OLMALI)
            if previous_y < line_bottom_y <= center_y_orig and tracker_id in entry_times:
                # Zamanı ve hızı hesapla
                entry_time = entry_times[tracker_id]
                exit_time = time.time()
                time_diff = exit_time - entry_time
                
                # Çok kısa sürede geçişler hatalı olabilir, onları ele
                if time_diff < 0.1: 
                    del entry_times[tracker_id]
                    continue

                speed_ms = distance_m / time_diff  # m/s
                speed_kmh = speed_ms * 3.6  # km/h
                
                print(f"--- KAMERA {video_index+1}: ID {tracker_id} ÇIKIŞ YAPTI. Hız: {int(speed_kmh)} km/h ---")

                if speed_kmh > HIZ_LIMITI_KMH:
                    # İhlal tespiti
                    # 'best_shot' mantığı için xyxy'yi orijinal frame'e ölçekle
                    xyxy_orig = xyxy * np.array([original_width / W, original_height / H, original_width / W, original_height / H])
                    best_shot_info[tracker_id] = {'frame': frame.copy(), 'xyxy': xyxy_orig}
                    final_violators.append({'tracker_id': tracker_id, 'hiz': int(speed_kmh), 'timestamp': datetime.now()})

                # Bu araçla işimiz bitti, listeden çıkaralım
                del entry_times[tracker_id]
        
        # --- DEĞİŞTİ: GÖRSELLEŞTİRME ---
        # İki çizgiyi de çiz
        annotated_frame = box_annotator.annotate(scene=low_res_frame.copy(), detections=detections)
        line_top_y_low_res = int(H * line_top_ratio)
        line_bottom_y_low_res = int(H * line_bottom_ratio)
        cv2.line(annotated_frame, (0, line_top_y_low_res), (W, line_top_y_low_res), (0, 255, 0), 2)
        cv2.line(annotated_frame, (0, line_bottom_y_low_res), (W, line_bottom_y_low_res), (0, 255, 0), 2)
        
        with frame_locks[video_index]:
            ret, buffer = cv2.imencode('.jpg', annotated_frame)
            output_frames[video_index] = buffer.tobytes()

    cap.release()
    print(f"Kamera {video_index+1} için hızlı tarama tamamlandı. {len(final_violators)} potansiyel ihlal bulundu.")
    
    # Yüksek kaliteli ihlal işleme ve raporlama kısmı önceki kod ile aynı.
    # Sadece best_shot'ı ihlal anında yakaladığımız için mantık biraz daha basit.
    # ... (Bu kısım buraya gelecek) ...
    thread_local_ihlaller = []
    for violator in final_violators:
        tracker_id = violator['tracker_id']
        best_info = best_shot_info.get(tracker_id)
        if not best_info: continue
        arac_resmi = sv.crop_image(image=best_info['frame'], xyxy=best_info['xyxy'])
        plaka_metni, iyilestirilmis_plaka_resmi = iyilestir_ve_plaka_oku(arac_resmi, plaka_model, upsampler)
        
        timestamp_str = violator['timestamp'].strftime("%Y%m%d_%H%M%S")
        arac_resim_yolu = os.path.join(CIKTI_KLASORU, f"cam{video_index+1}_arac_{tracker_id}_{timestamp_str}.jpg")
        plaka_resim_yolu = os.path.join(CIKTI_KLASORU, f"cam{video_index+1}_plaka_{tracker_id}_{timestamp_str}.jpg")
        
        cv2.imwrite(arac_resim_yolu, arac_resmi)
        if iyilestirilmis_plaka_resmi is not None:
            cv2.imwrite(plaka_resim_yolu, iyilestirilmis_plaka_resmi)
        else:
            plaka_resim_yolu = None
        
        thread_local_ihlaller.append({
            "kamera_id": video_index + 1,
            "timestamp": violator['timestamp'].strftime("%Y-%m-%d %H:%M:%S"),
            "tracker_id": tracker_id, "hiz_kmh": violator['hiz'],
            "plaka_metni": plaka_metni or "OKUNAMADI",
            "arac_resim_yolu": arac_resim_yolu, "plaka_resim_yolu": plaka_resim_yolu,
        })
    with all_kayitlari_lock:
        all_ihlal_kayitlari.extend(thread_local_ihlaller)
    finished_frame = np.zeros((TARGET_HEIGHT, int(TARGET_HEIGHT * 16/9), 3), dtype=np.uint8)
    cv2.putText(finished_frame, f"Kamera {video_index+1}: Bitti", (50, int(TARGET_HEIGHT/2)), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
    with frame_locks[video_index]:
        ret, buffer = cv2.imencode('.jpg', finished_frame)
        output_frames[video_index] = buffer.tobytes()

# Rapor yöneticisi ve Flask route'ları önceki kod ile aynı.
# ... (generate_frame_for_web, report_manager, index, video_feed fonksiyonları buraya gelecek) ...
def generate_frame_for_web(video_index):
    while True:
        with frame_locks[video_index]:
            frame = output_frames[video_index]
        if frame is None:
            time.sleep(0.05)
            continue
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
        time.sleep(1 / VIDEO_FPS)

def report_manager():
    for t in processing_threads:
        t.join()
    print("\n==============================================")
    print("Tüm video akışları tamamlandı.")
    print("Birleşik rapor oluşturuluyor...")
    excel_raporu_olustur(all_ihlal_kayitlari, EXCEL_RAPOR_ADI)
    print("Tüm işlemler tamamlandı. Program sonlandırılabilir.")
    print("==============================================")

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE, video_count=len(VIDEO_CONFIGS))

@app.route('/video_feed/<int:video_index>')
def video_feed(video_index):
    if video_index >= len(VIDEO_CONFIGS):
        return "Geçersiz video indeksi", 404
    return Response(generate_frame_for_web(video_index),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

# ==============================================================================
# --- 5. ANA ÇALIŞTIRMA BLOĞU ---
# ==============================================================================
if __name__ == "__main__":
    os.makedirs(CIKTI_KLASORU, exist_ok=True)
    
    for i, config in enumerate(VIDEO_CONFIGS):
        thread = threading.Thread(target=process_video_stream, args=(i, config))
        thread.daemon = True
        processing_threads.append(thread)
        thread.start()

    report_thread = threading.Thread(target=report_manager)
    report_thread.daemon = True
    report_thread.start()

    print("\nWeb sunucusu başlatılıyor...")
    print(f"Arayüze erişmek için tarayıcınızda http://127.0.0.1:5000 adresini açın.")
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)